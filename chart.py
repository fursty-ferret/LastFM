from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

def main():
    filename = 'TopArtists.xlsx'
    wb = load_workbook(filename)
    with open('chart.json','r') as file:
        data = json.load(file)
    list = []
    for keys, values in enumerate(data):
        points = data[values]['total']
        firsts = data[values]['first']
        seconds = data[values]['second']
        thirds = data[values]['third']
        artist = values
        ls = [artist, firsts, seconds, thirds, points]
        list.append(ls)
    for x in range(len(list)):
        write_to_excel(list[x],wb)
    wb.save(filename)

def write_to_excel(list, wb):
    ws = wb['Overall Charts']
    row = str(ws.max_row +1)
    for i in range(0,5):
        column = str(get_column_letter(i+1))
        cell = column + row
        ws[cell] = list[i]


if __name__ == "__main__":
    main()
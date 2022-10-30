from Auth import *
import requests, math, csv, os, json
from UnixToDate import convert_unix
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

def main():
    for weeknumber in range(355,924):
    # weeknumber = int(input("What week to input? 919 = 02/10/22 "))
        unix = get_from_to(weeknumber)
        date_from = unix[0]
        date_to = unix[1]
        artist_url = root + '/2.0/?method=user.getweeklyartistchart&user=' + user + '&api_key=' + API_key + '&format=json' + '&from=' + str(date_from) + '&to=' + str(date_to)
        r = requests.get(artist_url)
        rjson = r.json()
        weeklyartistchart = rjson['weeklyartistchart']
        artist = weeklyartistchart['artist']
        date = convert_unix(date_to)
        with open('chart.json','r') as f:
            artistchart = json.load(f)
        print(f'Starting {weeknumber}, week ending {date}')
        if len(artist) < 3:
            print("Not enough artists recorded this week")
        else:
            total = 0
            for l in range(len(artist)):
                playcount = artist[l]['playcount']
                total += int(playcount)
            unique = len(artist)
            weeklychartlist = [date, total, unique]
            for i in range(0,3):
                    name = artist[i]['name']
                    playcount = int(artist[i]['playcount'])
                    percentage = int(playcount)/total
                    results = [name, playcount, percentage]
                    weeklychartlist.extend(results)
                    while name not in artistchart:
                        artistchart.update({name:{'first':0,'second':0,'third':0,'total':0}})
                    else:
                        artistchart[name]['total'] += 3-i
                        if i == 0:
                            artistchart[name]['first'] += 1
                        elif i == 1:
                            artistchart[name]['second'] += 1
                        elif i == 2:
                            artistchart[name]['third'] += 1
            with open('chart.json','w') as file:
                json.dump(artistchart, file, ensure_ascii=True)
            print(f'Written {weeknumber}')
            # write_to_excel(weeklychartlist)
            print(f'Complete {date}')

def write_to_excel(weeklychartlist):
    filename = 'TopArtists.xlsx'
    wb = load_workbook(filename)
    ws = wb['Weekly Charts']
    row = str(ws.max_row + 1)
    fontObj = Font(bold=True)
    aborder = Border(left=Side(border_style='dashed'), right=Side(border_style='medium'), bottom=Side(border_style='dashed'), top=Side(border_style='dashed'))
    cellborder = Border(left=Side(border_style='dashed'), right=Side(border_style='dashed'), bottom=Side(border_style='dashed'), top=Side(border_style='dashed'))
    ws.row_dimensions[int(row)].height = 25
    column_widths = [15, 8, 8, 30, 5, 5, 30, 5, 5, 30, 5, 5]
    for cells in range(0,12):
        column = str(get_column_letter(cells+1))
        ws.column_dimensions[column].width = column_widths[cells]
        cell = column + row
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        ws[cell].border = cellborder
        ws[cell] =  weeklychartlist[cells]
    ws['A'+row].font, ws['A'+row].border = fontObj, aborder
    wb.save(filename)

def get_from_to(count):
    chart_url = root + '/2.0/?method=user.getweeklychartlist&user=' + user + '&api_key=' + API_key + '&format=json'
    r = requests.get(chart_url)
    rjson = r.json()
    weeklychartlist = rjson['weeklychartlist']
    chart = weeklychartlist['chart']
    date = chart[count]
    funix = int(date['from'])
    tunix = int(date['to']) 
    return funix, tunix

if __name__ == '__main__':
    main()